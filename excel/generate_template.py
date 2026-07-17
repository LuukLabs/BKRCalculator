#!/usr/bin/env python3
"""
Generates BKR-rekentool.xlsx — an Excel template that reproduces the BKRCalculator
engine (KDVManager.BKRCalculator.GroupAnalyzer) entirely with worksheet formulas.

The rule table below is a 1:1 copy of AgeGroupRulesFactory.BuildAgeGroupRules(). Keep
the two in sync: re-run this script if the rules change. The formula logic was verified
against the C# engine across all 83,521 child-count combinations (0..16 per age) with
zero mismatches.

Requires: Excel 2021 or Microsoft 365 (uses dynamic-array evaluation inside MATCH).
Run: python3 generate_template.py
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.workbook.defined_name import DefinedName

# (MinAge, MaxAge, MinProfessionals, MaxChildren, MaxAge0) — MaxAge0 caps the 0-1 sub-group; 999 = no cap.
RULES = [
    (0, 1, 1, 3, 999), (0, 1, 2, 6, 999), (0, 1, 3, 9, 999), (0, 1, 4, 12, 999),
    (1, 2, 1, 5, 999), (1, 2, 2, 10, 999), (1, 2, 3, 15, 999), (1, 2, 4, 16, 999),
    (2, 3, 1, 8, 999), (2, 3, 2, 16, 999),
    (3, 4, 1, 8, 999), (3, 4, 2, 16, 999),
    (0, 2, 1, 4, 999), (0, 2, 2, 8, 999), (0, 2, 3, 14, 8), (0, 2, 4, 16, 8),
    (0, 3, 1, 5, 999), (0, 3, 2, 10, 999), (0, 3, 3, 13, 8), (0, 3, 3, 14, 4),
    (0, 3, 3, 15, 3), (0, 3, 4, 16, 8),
    (0, 4, 1, 5, 999), (0, 4, 2, 12, 999), (0, 4, 3, 14, 8), (0, 4, 3, 15, 5),
    (0, 4, 3, 16, 3), (0, 4, 4, 16, 8),
    (1, 3, 1, 6, 999), (1, 3, 2, 11, 999), (1, 3, 3, 16, 999),
    (1, 4, 1, 7, 999), (1, 4, 2, 13, 999), (1, 4, 3, 16, 999),
    (2, 4, 1, 8, 999), (2, 4, 2, 16, 999),
]

ACCENT = "2F5496"
HEADER_FILL = PatternFill("solid", fgColor=ACCENT)
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
RESULT_FILL = PatternFill("solid", fgColor="E2EFDA")
WHITE_BOLD = Font(bold=True, color="FFFFFF")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def build():
    wb = Workbook()

    # ---- Sheet: Rekentool (user-facing) ----
    ws = wb.active
    ws.title = "Rekentool"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 16

    ws["A1"] = "BKR-rekentool — beroepskracht-kindratio"
    ws["A1"].font = Font(bold=True, size=14, color=ACCENT)
    ws["A2"] = "Vul het aantal kinderen per leeftijd in. Het minimum aantal beroepskrachten verschijnt onderaan."
    ws["A2"].font = Font(italic=True, size=9, color="808080")

    ws["A3"] = "Aantal kinderen per leeftijd"
    ws["A3"].font = WHITE_BOLD
    ws["A3"].fill = HEADER_FILL
    ws["B3"].fill = HEADER_FILL
    labels = ["0 jaar (0–1)", "1 jaar (1–2)", "2 jaar (2–3)", "3 jaar (3–4)"]
    for i, label in enumerate(labels):
        row = 4 + i
        ws[f"A{row}"] = label
        cell = ws[f"B{row}"]
        cell.value = 0
        cell.fill = INPUT_FILL
        cell.border = BOX
        cell.alignment = Alignment(horizontal="center")

    ws["A9"] = "Benodigde beroepskrachten:"
    ws["A9"].font = Font(bold=True, size=12)
    ws["A9"].fill = RESULT_FILL
    res = ws["B9"]
    res.value = (
        '=IF(SUM($B$4:$B$7)=0,0,'
        'IF(Berekening!M2=0,"Ongeldige groepssamenstelling",'
        'Berekening!M2+IF(MAX(Berekening!M3:M6)>Berekening!M2,1,0)))'
    )
    res.font = Font(bold=True, size=12, color=ACCENT)
    res.fill = RESULT_FILL
    res.alignment = Alignment(horizontal="center")
    res.border = BOX

    notes = [
        "",
        "Toelichting",
        "• Ratio per leeftijd: 0 jr 1:3 · 1 jr 1:5 · 2 jr 1:6 · 3 jr 1:8 (dagopvang, vanaf 1 juli 2024).",
        "• De uitkomst houdt rekening met de maximale groepsgrootte én met de extra-kracht-regel",
        "  (één kind minder mag nooit meer beroepskrachten vereisen).",
        '• "Ongeldige groepssamenstelling" betekent dat geen wettelijke groepsnorm op deze',
        "  combinatie past (bijv. te grote of niet-toegestane mengvorm).",
        "• Berekend volgens dezelfde regels als de BKRCalculator-engine; bedoeld als hulpmiddel,",
        "  de officiële rekentool op 1ratio.nl is bindend.",
    ]
    for i, text in enumerate(notes):
        r = 11 + i
        ws[f"A{r}"] = text
        if text == "Toelichting":
            ws[f"A{r}"].font = Font(bold=True)
        else:
            ws[f"A{r}"].font = Font(size=9, color="606060")

    # ---- Sheet: Regels (the legal norm table) ----
    rules = wb.create_sheet("Regels")
    headers = ["MinLeeftijd", "MaxLeeftijd", "MinBeroepskrachten", "MaxKinderen", "MaxNuljarigen"]
    for c, h in enumerate(headers, start=1):
        cell = rules.cell(row=1, column=c, value=h)
        cell.font = WHITE_BOLD
        cell.fill = HEADER_FILL
    for r, rule in enumerate(RULES, start=2):
        for c, val in enumerate(rule, start=1):
            rules.cell(row=r, column=c, value=val)
    for col, w in zip("ABCDE", (12, 12, 20, 13, 15)):
        rules.column_dimensions[col].width = w

    last = 1 + len(RULES)  # row 37
    wb.defined_names.add(DefinedName("MinA", attr_text=f"Regels!$A$2:$A${last}"))
    wb.defined_names.add(DefinedName("MaxA", attr_text=f"Regels!$B$2:$B${last}"))
    wb.defined_names.add(DefinedName("MinP", attr_text=f"Regels!$C$2:$C${last}"))
    wb.defined_names.add(DefinedName("MaxC", attr_text=f"Regels!$D$2:$D${last}"))
    wb.defined_names.add(DefinedName("MaxC0", attr_text=f"Regels!$E$2:$E${last}"))

    # ---- Sheet: Berekening (helper; mirrors GroupAnalyzer step by step) ----
    calc = wb.create_sheet("Berekening")
    calc.sheet_state = "hidden"
    hdr = ["Scenario", "x0", "x1", "x2", "x3", "totaal", "minLft", "maxLft",
           "regel#", "minBK", "ratioBK", "basis", "telwaarde"]
    for c, h in enumerate(hdr, start=1):
        calc.cell(row=1, column=c, value=h).font = Font(bold=True)

    # Five scenarios: original group, then the group with one child fewer per present age.
    # x-values per scenario (Excel refs to the Rekentool inputs).
    I0, I1, I2, I3 = "Rekentool!$B$4", "Rekentool!$B$5", "Rekentool!$B$6", "Rekentool!$B$7"
    scenarios = [
        ("Origineel",        I0,            I1,            I2,            I3),
        ("Min 1× 0-jarige",  f"{I0}-1",     I1,            I2,            I3),
        ("Min 1× 1-jarige",  I0,            f"{I1}-1",     I2,            I3),
        ("Min 1× 2-jarige",  I0,            I1,            f"{I2}-1",     I3),
        ("Min 1× 3-jarige",  I0,            I1,            I2,            f"{I3}-1"),
    ]
    # telwaarde guard: a reduced scenario only counts if that age is actually present.
    present_guard = [None, I0, I1, I2, I3]

    for i, (label, x0, x1, x2, x3) in enumerate(scenarios):
        r = 2 + i
        calc.cell(row=r, column=1, value=label)
        calc.cell(row=r, column=2, value=f"={x0}")
        calc.cell(row=r, column=3, value=f"={x1}")
        calc.cell(row=r, column=4, value=f"={x2}")
        calc.cell(row=r, column=5, value=f"={x3}")
        calc.cell(row=r, column=6, value=f"=B{r}+C{r}+D{r}+E{r}")  # totaal
        calc.cell(row=r, column=7,
                  value=f"=IF(B{r}>0,0,IF(C{r}>0,1,IF(D{r}>0,2,IF(E{r}>0,3,-1))))")  # minLft
        calc.cell(row=r, column=8,
                  value=f"=IF(E{r}>0,3,IF(D{r}>0,2,IF(C{r}>0,1,IF(B{r}>0,0,-1))))")  # maxLft
        calc.cell(row=r, column=9,
                  value=(f"=IFERROR(MATCH(1,(MinA=G{r})*(MaxA=H{r}+1)*(MaxC>=F{r})"
                         f"*(MaxC0>=B{r}),0),0)"))  # regel#
        calc.cell(row=r, column=10, value=f"=IF(I{r}=0,0,INDEX(MinP,I{r}))")  # minBK
        calc.cell(row=r, column=11,
                  value=f"=IF(B{r}>0,CEILING(B{r}/3+(C{r}/5+D{r}/6+E{r}/8)/1.2,1),0)")  # ratioBK
        calc.cell(row=r, column=12, value=f"=IF(OR(F{r}=0,I{r}=0),0,MAX(J{r},K{r}))")  # basis
        if i == 0:
            calc.cell(row=r, column=13, value=f"=L{r}")  # telwaarde (original)
        else:
            calc.cell(row=r, column=13, value=f"=IF({present_guard[i]}>0,L{r},0)")

    out = "BKR-rekentool.xlsx"
    wb.save(out)
    print(f"wrote {out} ({len(RULES)} rules)")


if __name__ == "__main__":
    build()
